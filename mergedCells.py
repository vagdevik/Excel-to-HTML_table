#####################
#
# xlrd -  to read and extract info regarding cell
# openpyxl - to extract merged cells
# ------------------
# to run:
# python merdegCells.py <path to input excel file>
# <path to input excel file> : '/home/vubuntu/Documents/Others/Intern/Great Four/excel_file.xlsx'
#
#####################


from openpyxl.utils import coordinate_from_string,column_index_from_string
import xlrd
import xlsxwriter
from datetime import datetime
from openpyxl import load_workbook,styles
import sys
reload(sys)


sys.setdefaultencoding('utf8')
## arg[0] - code file
## arg[1] - input file name
input_filename = sys.argv[1]
excel_file = xlrd.open_workbook(input_filename,on_demand=True)
sheetnames = excel_file.sheet_names()

def floatHourToTime(fh):
    '''
    converts xltime to python datetime mode
    source: https://stackoverflow.com/questions/31359150/convert-date-from-excel-in-number-format-to-date-format-python
    '''
    h, r = divmod(fh, 1)
    m, r = divmod(r*60, 1)
    return (int(h),int(m),int(r*60))





head_cells={} # head_cells["(1,2)":"0"], this "0" is the key in cell_dict
cell_dict={}  # contains info regarding colspan,rowspan of each cell, cell_dict["0"]={"colspan"="3",rowspan":"2"}
hidden_cells=[] # list of all the hidden cells including head cells

def mergedcell_info(merged_ranges):
	'''
	returns head merged cells and hidden cells under the merged ones
	'''
	
	global cell_dict
	global head_cells
	global hidden_cells
	ind=0
	for mrange in merged_ranges:
		'''
		to convert merged range, say A1:B3 to (0,0) to (1,2)
		'''
		
		mrange = str(mrange)
		cell = mrange.split(':')

		xy = coordinate_from_string(cell[0])
		col1 = column_index_from_string(xy[0])
		row1 = xy[1]

		head_cells[str((row1-1,col1-1))] = str(ind)
		

		xy = coordinate_from_string(cell[1])
		col2 = column_index_from_string(xy[0])
		row2 = xy[1]

		for i in range(row1-1,row2):
			for j in range(col1-1,col2):
				hidden_cells.append((i,j))
	
		rowspan = (row2-row1)+1
		colspan = (col2-col1)+1
		
		cell_dict[str(ind)] = {}
		cell_dict[str(ind)]["rowspan"] = str(rowspan) 
		cell_dict[str(ind)]["colspan"] = str(colspan)

		ind+=1
		
def get_cell_value(sheet,i,j):
	'''
	extracts the exact cell value after necessary conversions, say, xldate to dattime
	'''
	cell_type = sheet.cell(i,j).ctype
	cell_value = sheet.cell(i,j).value
	if cell_type == 3: #converts xldate format to datetime format
		excel_date = cell_value
		dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(excel_date) - 2)
		hour, minute, second = floatHourToTime(excel_date % 1)
		dt = dt.replace(hour=hour, minute=minute, second=second)
		value = str(dt.month)+"/"+str(dt.day)+"/"+str(dt.year) #date format : month/day/year
	elif cell_type == 2:
		deci_nums=[]
		col = sheet.col(j)
		count=0
		for element in col:
			if element.ctype == 2:
				count += 1
				num_list = str(element.value).split('.')
				deci_nums.append(str(num_list[1]))
		if (len(set(deci_nums))==1) and deci_nums[0]=="0":
			value = str(sheet.cell(i, j).value).split('.')[0]
	else:
		value = str(cell_value)
	return value		 
		
def get_cell_bgcolor(sheet,i,j):
	'''
	returns background color of the cell
	fgColor sources: https://stackoverflow.com/questions/40471653/how-to-select-cell-with-specific-color-in-openpyxl
	xl_col_to_name source: https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
	font_color source : https://stackoverflow.com/questions/30483649/python-openpyxl-how-to-read-a-cell-font-color
	'''

	cell_index = xlsxwriter.utility.xl_col_to_name(j)+str(i+1) 

	bcolor = sheet[cell_index].fill.fgColor
	bgcolor = '#'+str(bcolor.rgb)[2:]
	
	fcolor = sheet[cell_index].font.color
	print fcolor
	try:
		font_color = '#'+str(fcolor.rgb)[2:]
	except:
		font_color = '#000000'	
	return str(bgcolor),str(font_color)



with open('html_table.html','w') as f:
	f.write('<!DOCTYPE html>\n<html>\n')
	f.write('<head>\n')
	f.write('<meta http-equiv="content-type" content="text/html; charset=UTF-8">\n')
	f.write('<script type="text/javascript" src="https://code.jquery.com/jquery-2.1.3.js"></script>\n')
	#f.write('<link rel="stylesheet" type="text/css" href="/css/result-light.css">\n')
	f.write('<link rel="stylesheet" type="text/css" href="styles.css">\n')
	f.write('<script type="text/javascript" src="myscripts.js"></script>\n')
	f.write('</head>\n')
	f.write('<body>\n')
	for i in range(len(sheetnames)):
		f.write('<br>\n')
		f.write("<form>\n<input type='button' value='"+str(sheetnames[i])+ "' onclick='sheet_data("+"'"+sheetnames[i]+"'"+")'>\n</form>\n")
	f.write('<br>\n')
	f.write('<center>\n<button onclick="myFunction()">Form a mini table</button>\n</center>\n<br>\n')
	f.write("<div id = 'selectedTable'></div>\n</center>\n<br><br>\n")
	sheet = excel_file.sheet_by_index(0) #sheet of 0th index in thw excel workbook
	nrows = sheet.nrows #no of rows in sheet
	ncols = sheet.ncols #no of cols in sheet
	
	workbook_openpyxl = load_workbook(filename = input_filename,data_only=True)
	sheet_openpyxl = workbook_openpyxl.worksheets[0]
	merged_ranges = sheet_openpyxl.merged_cell_ranges #get merged cells in a list

	# get the cells under the merged one
	mergedcell_info(merged_ranges)
	f.write('<div id = "myDiv">\n')
	f.write("<table border='1'>\n")
	for i in range(nrows):
		f.write("<tr rowid = '"+ str(i) +"'>\n")
		for j in range(ncols):
			value = get_cell_value(sheet,i,j)
			bgcolor,font_color = get_cell_bgcolor(sheet_openpyxl,i,j)
			if bgcolor == '#000000':
				bgcolor = '#ffffff'
			if (i,j) not in hidden_cells:
				f.write("<td rowid = '"+ str(i) +"' colid = '"+ str(j) +"' bgcolor = '"+bgcolor+"' style = 'color:"+font_color+";'>"+value+"</td>\n")
			elif str((i,j)) in head_cells:
				f.write("<td rowid = '"+ str(i) +"' colid = '"+ str(j) +"' bgcolor = '"+bgcolor+"' colspan = '"+cell_dict[head_cells[str((i,j))]]["colspan"]+"' rowspan = '"+cell_dict[head_cells[str((i,j))]]["rowspan"]+"' style = 'color:"+font_color+";'>"+value+"</td>\n")
		f.write("\n</tr>\n")	
	f.write("</table>\n")
	f.write("</div>\n")
	#f.write('<script>\nfunction myFunction() {\nhtml = "";\nvar t = $(".selected");\nif (t.length == 0){\nalert("Please select your cells from the table first.");\n}\nfor (var i=0;i<t.length;i++){\nconsole.log("******");\n	console.log($(".selected")[i].outerHTML);\nconsole.log("@@@@@@");\nhtml = html + $(".selected")[i].outerHTML;\n}\n	console.log("-----");\n	console.log(html);\n}\n</script>')
	
	f.write("<script>\n")
	f.write("function myFunction() {\n")
	f.write("html = '';\n")
	f.write("var t = $('.selected');\n")
	f.write("if (t.length == 0){\n")
	f.write("alert('Please select your cells from the table first.');\n")
	f.write("}\n")
	f.write("else{\n")
	f.write("htmlCode = ''\n")
	f.write("var prevRow = parseInt($('.selected')[0].getAttribute('rowid'));\n")
	f.write("htmlCode = htmlCode + \"<table border = " + "'1'>" + "<tr " + "rowid = '\"" + " + prevRow + " + "\"'>\"\n")
	f.write("for (var i=0;i<t.length;i++){\n")
	f.write("var curr = $('.selected')[i];\n")
	f.write("var currRow = parseInt(curr.getAttribute('rowid'));\n")
	f.write("if(currRow>prevRow){\n")
	f.write("prevRow = currRow;\n")
	f.write("htmlCode = htmlCode + \"</tr><tr \"+\"rowid = '\"+prevRow+\"'>\";\n")
	f.write("}\n")
	f.write("html = html + curr.outerHTML;\n")
	f.write("htmlCode = htmlCode + curr.outerHTML;\n")
	f.write("}\n")
	f.write("htmlCode = htmlCode + '</table>';\n")
	f.write("}\n")
	f.write("console.log(htmlCode);\n")
	f.write("string = '<p>You selected the following table</p>';\n")
	f.write("document.getElementById('selectedTable').innerHTML = string + htmlCode;\n")
	f.write("$('.selected').removeClass('selected');")
	f.write("}\n")
	f.write("</script>\n")

	f.write("</body>\n")
	f.write("</html>\n")
	f.close()
